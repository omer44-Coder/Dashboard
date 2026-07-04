import openpyxl, json, re, glob, sys
from datetime import datetime

SHEET_NAMES = ['Genel','Genel Müşteri Çıkış','Adresleme','Toplama','Sorter','Tekli-Eksik Sevk','Kargo','Adres Denetim','İade Kabul']
AY_ISIMLERI = ['Ocak','Şubat','Mart','Nisan','Mayıs','Haziran','Temmuz','Ağustos','Eylül','Ekim','Kasım','Aralık']

xlsx_files = glob.glob('*.xlsx')
if not xlsx_files:
    print("HATA: Repo icinde .xlsx dosyasi bulunamadi.")
    sys.exit(1)
xlsx_path = xlsx_files[0]
print("Excel dosyasi bulundu: " + xlsx_path)

wb = openpyxl.load_workbook(xlsx_path, data_only=True)

def to_ymd(d):
    return str(d.year).zfill(4) + "-" + str(d.month).zfill(2) + "-" + str(d.day).zfill(2)

newData = {}

for sName in SHEET_NAMES:
    if sName not in wb.sheetnames:
        continue
    ws = wb[sName]
    rows = list(ws.iter_rows(values_only=True))
    n = len(rows)
    blocks = []
    for i in range(n):
        row = rows[i]
        if row and len(row) > 1 and row[1] == 'Tarih':
            dates = [v for v in row[2:] if isinstance(v, datetime)]
            if dates:
                year_counts = {}
                for d in dates:
                    year_counts[d.year] = year_counts.get(d.year, 0) + 1
                majority_year = max(year_counts, key=year_counts.get)
                blocks.append({'dateRow': i, 'haftaRow': i+1, 'saatRow': i+2, 'adetRow': i+3, 'year': majority_year})
    newData[sName] = {}
    for b in blocks:
        yr = b['year']
        sfx = str(yr)[-2:]
        date_row = rows[b['dateRow']]
        haftaRowIdx = b['haftaRow']
        saatRowIdx = b['saatRow']
        adetRowIdx = b['adetRow']
        hafta_row = rows[haftaRowIdx] if haftaRowIdx < n else []
        saat_row = rows[saatRowIdx] if saatRowIdx < n else []
        adet_row = rows[adetRowIdx] if adetRowIdx < n else []
        arr = []
        for c in range(2, len(date_row)):
            raw = date_row[c]
            if raw is None or not isinstance(raw, datetime):
                continue
            d = raw
            if d.year != yr:
                continue
            saat = None
            adet = None
            if c < len(saat_row) and saat_row[c] is not None:
                try:
                    saat = float(saat_row[c])
                except (ValueError, TypeError):
                    saat = None
            if c < len(adet_row) and adet_row[c] is not None:
                try:
                    adet = float(adet_row[c])
                except (ValueError, TypeError):
                    adet = None
            if not saat or not adet or saat <= 0 or adet <= 0:
                continue
            hafta = hafta_row[c] if c < len(hafta_row) and hafta_row[c] is not None else ''
            m = re.search(r'\d+', str(hafta))
            hno = int(m.group()) if m else None
            arr.append({'ymd': to_ymd(d), 'ay': d.month, 'hno': hno, 'saat': saat, 'adet': adet})
        gun = {}
        for r in arr:
            gun[r['ymd']] = {'saat': round(r['saat']*100)/100, 'adet': round(r['adet'])}
        hft = {}
        for r in arr:
            if not r['hno']:
                continue
            k = str(r['hno']) + '. Hafta'
            if k not in hft:
                hft[k] = {'saat': 0, 'adet': 0}
            hft[k]['saat'] += r['saat']
            hft[k]['adet'] += r['adet']
        hft_sorted = dict(sorted(hft.items(), key=lambda x: int(x[0].split('.')[0])))
        ay = {}
        for r in arr:
            k = AY_ISIMLERI[r['ay']-1]
            if k not in ay:
                ay[k] = {'saat': 0, 'adet': 0}
            ay[k]['saat'] += r['saat']
            ay[k]['adet'] += r['adet']
        ts = None
        if saat_row and saat_row[0] is not None:
            try:
                ts = float(saat_row[0])
            except (ValueError, TypeError):
                ts = None
        if not ts:
            ts = sum(r['saat'] for r in arr)
        ta = None
        if adet_row and adet_row[0] is not None:
            try:
                ta = float(adet_row[0])
            except (ValueError, TypeError):
                ta = None
        if not ta:
            ta = sum(r['adet'] for r in arr)
        newData[sName]['gun' + sfx] = gun
        newData[sName]['hft' + sfx] = hft_sorted
        newData[sName]['ay' + sfx] = ay
        newData[sName]['tot' + sfx] = {'saat': round(ts*100)/100, 'adet': round(ta)}

new_data_str = json.dumps(newData, ensure_ascii=False, separators=(',', ':'))

with open('index.html', 'r', encoding='utf-8') as f:
    html = f.read()

m = re.search(r'let DATA = (\{.*?\});\nlet LAST_UPDATE_DATA', html, re.DOTALL)
if not m:
    print("HATA: index.html icinde DATA blogu bulunamadi.")
    sys.exit(1)

new_html = html[:m.start(1)] + new_data_str + html[m.end(1):]

now_str = datetime.now().strftime('%Y-%m-%dT%H:%M:%S')
new_html = re.sub(r'let LAST_UPDATE_DATA = "[^"]*";', 'let LAST_UPDATE_DATA = "' + now_str + '";', new_html)
new_html = re.sub(r'let LAST_UPDATE = "[^"]*";', 'let LAST_UPDATE = "' + now_str + '";', new_html)

with open('index.html', 'w', encoding='utf-8') as f:
    f.write(new_html)

print("index.html basariyla guncellendi.")
